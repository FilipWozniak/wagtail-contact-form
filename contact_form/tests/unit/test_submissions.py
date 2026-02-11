from datetime import datetime
from io import BytesIO

import pytest
from django.contrib.auth import get_user_model
from django.test import Client
from django.urls import reverse
from wagtail.contrib.forms.models import FormSubmission
from wagtail.models import Site

from contact_form.models import ContactPage
from contact_form.models import FormField
from contact_form.views import CustomFormPagesListView
from contact_form.views import CustomSubmissionsListView
from contact_form.views import FormPagePaginator
from contact_form.views import FormPagesFilterSet
from contact_form.views import FormSubmissionPaginator

User = get_user_model()


@pytest.fixture
def admin_user(db):
    return User.objects.create_superuser(
        username="admin",
        email="admin@example.com",
        password="admin_password",
    )


@pytest.fixture
def admin_client(admin_user):
    client = Client()
    client.force_login(admin_user)
    return client


@pytest.fixture
def contact_page(db):
    page = ContactPage(
        title="Contact Us",
        intro="Lorem Ipsum",
        thank_you_text="Thank You",
        from_address="from@example.com",
        to_address="to@example.com",
        subject="Contact Form",
    )
    home_page = Site.objects.filter(is_default_site=True)[0].root_page
    home_page.add_child(instance=page)
    page.save()

    FormField.objects.create(
        page=page,
        sort_order=1,
        label="Full Name",
        field_type="singleline",
        required=True,
    )
    FormField.objects.create(
        page=page,
        sort_order=2,
        label="E-Mail Address",
        field_type="email",
        required=True,
    )
    FormField.objects.create(
        page=page,
        sort_order=3,
        label="Message",
        field_type="multiline",
        required=True,
    )

    return page


@pytest.fixture
def form_submissions(contact_page):
    submissions = []
    for i in range(15):
        submission = FormSubmission.objects.create(
            page=contact_page,
            form_data={
                "full_name": f"User {i}",
                "e-mail_address": f"user{i}@example.com",
                "message": f"Message {i}" * 20,
            },
        )
        submissions.append(submission)
    return submissions


@pytest.mark.django_db
class TestCustomSubmissionsListView:
    def test_pagination_setting(self):
        assert CustomSubmissionsListView.paginate_by == 10

    def test_page_title(self):
        assert CustomSubmissionsListView.page_title == "Form Data"

    def test_paginator_class(self):
        assert CustomSubmissionsListView.paginator_class == FormSubmissionPaginator

    def test_paginator_verbose_name_plural(self):
        paginator = FormSubmissionPaginator([], per_page=10)
        assert paginator.verbose_name_plural == "Form Submissions"

    def test_paginator_verbose_name(self):
        paginator = FormSubmissionPaginator([], per_page=10)
        assert paginator.verbose_name == "Form Submission"


@pytest.mark.django_db
class TestSubmissionsListViewFormatting:
    def test_message_truncation(self, admin_client, contact_page):
        long_message = "A" * 200
        FormSubmission.objects.create(
            page=contact_page,
            form_data={
                "full_name": "Test User",
                "e-mail_address": "test@example.com",
                "message": long_message,
            },
        )
        url = reverse("custom_contact_form:list_submissions", args=[contact_page.pk])
        response = admin_client.get(url)
        content = response.content.decode("utf-8")
        assert "(...)" in content
        assert long_message not in content

    def test_date_formatting(self, admin_client, contact_page):
        submission = FormSubmission.objects.create(
            page=contact_page,
            form_data={
                "full_name": "Test User",
                "e-mail_address": "test@example.com",
                "message": "Test Message",
            },
        )
        submission.submit_time = datetime(2025, 1, 12, 9, 45)
        submission.save()
        url = reverse("custom_contact_form:list_submissions", args=[contact_page.pk])
        response = admin_client.get(url)
        content = response.content.decode("utf-8")
        assert "12 January 2025 at 09:45" in content

    def test_submission_date_column_is_last(self, admin_client, contact_page):
        FormSubmission.objects.create(
            page=contact_page,
            form_data={
                "full_name": "Test User",
                "e-mail_address": "test@example.com",
                "message": "Test Message",
            },
        )
        url = reverse("custom_contact_form:list_submissions", args=[contact_page.pk])
        response = admin_client.get(url)
        content = response.content.decode("utf-8")
        listing_start = content.find('class="listing"')
        table_section = content[listing_start : listing_start + 2000]
        full_name_pos = table_section.find("Full Name")
        submission_date_pos = table_section.find("Submission Date")
        assert full_name_pos > 0
        assert submission_date_pos > 0
        assert submission_date_pos > full_name_pos

    def test_submission_date_label(self, admin_client, contact_page):
        FormSubmission.objects.create(
            page=contact_page,
            form_data={
                "full_name": "Test User",
                "e-mail_address": "test@example.com",
                "message": "Test Message",
            },
        )
        url = reverse("custom_contact_form:list_submissions", args=[contact_page.pk])
        response = admin_client.get(url)
        content = response.content.decode("utf-8")
        assert "Submission Date" in content


@pytest.mark.django_db
class TestCSVDownload:
    def test_csv_download_returns_200(self, admin_client, contact_page, form_submissions):
        url = reverse(
            "custom_contact_form:list_submissions",
            args=[contact_page.pk],
        )
        response = admin_client.get(url, {"export": "csv"})
        assert response.status_code == 200
        assert "text/csv" in response["Content-Type"]

    def test_csv_download_contains_headers(self, admin_client, contact_page, form_submissions):
        url = reverse(
            "custom_contact_form:list_submissions",
            args=[contact_page.pk],
        )
        response = admin_client.get(url, {"export": "csv"})
        content = b"".join(response.streaming_content).decode("utf-8-sig")
        assert "Full Name" in content or "full_name" in content or "Submission" in content

    def test_csv_download_contains_data(self, admin_client, contact_page, form_submissions):
        url = reverse(
            "custom_contact_form:list_submissions",
            args=[contact_page.pk],
        )
        response = admin_client.get(url, {"export": "csv"})
        content = b"".join(response.streaming_content).decode("utf-8-sig")
        assert "User" in content

    def test_csv_download_has_correct_filename(self, admin_client, contact_page, form_submissions):
        url = reverse(
            "custom_contact_form:list_submissions",
            args=[contact_page.pk],
        )
        response = admin_client.get(url, {"export": "csv"})
        assert "attachment" in response["Content-Disposition"]
        assert ".csv" in response["Content-Disposition"]


@pytest.mark.django_db
class TestXLSXDownload:
    def test_xlsx_download_returns_200(self, admin_client, contact_page, form_submissions):
        url = reverse(
            "custom_contact_form:list_submissions",
            args=[contact_page.pk],
        )
        response = admin_client.get(url, {"export": "xlsx"})
        assert response.status_code == 200
        content_type = response["Content-Type"]
        assert "spreadsheetml" in content_type or "xlsx" in content_type

    def test_xlsx_download_has_correct_filename(self, admin_client, contact_page, form_submissions):
        url = reverse(
            "custom_contact_form:list_submissions",
            args=[contact_page.pk],
        )
        response = admin_client.get(url, {"export": "xlsx"})
        assert "attachment" in response["Content-Disposition"]
        assert ".xlsx" in response["Content-Disposition"]

    def test_xlsx_download_is_valid_file(self, admin_client, contact_page, form_submissions):
        try:
            import openpyxl
        except ImportError:
            pytest.skip("Package openpyxl is Not Installed")

        url = reverse(
            "custom_contact_form:list_submissions",
            args=[contact_page.pk],
        )
        response = admin_client.get(url, {"export": "xlsx"})
        content = b"".join(response.streaming_content)
        workbook = openpyxl.load_workbook(BytesIO(content))
        assert workbook.active is not None

    def test_xlsx_download_contains_data(self, admin_client, contact_page, form_submissions):
        try:
            import openpyxl
        except ImportError:
            pytest.skip("Package openpyxl is Not Installed")

        url = reverse(
            "custom_contact_form:list_submissions",
            args=[contact_page.pk],
        )
        response = admin_client.get(url, {"export": "xlsx"})
        content = b"".join(response.streaming_content)
        workbook = openpyxl.load_workbook(BytesIO(content))
        sheet = workbook.active
        values = []
        for row in sheet.iter_rows(values_only=True):
            values.extend([str(cell) for cell in row if cell])
        all_values = " ".join(values)
        assert "User" in all_values


@pytest.mark.django_db
class TestSubmissionsListView:
    def test_list_submissions_view_returns_200(self, admin_client, contact_page, form_submissions):
        url = reverse(
            "custom_contact_form:list_submissions",
            args=[contact_page.pk],
        )
        response = admin_client.get(url)
        assert response.status_code == 200

    def test_pagination_limits_results(self, admin_client, contact_page, form_submissions):
        url = reverse(
            "custom_contact_form:list_submissions",
            args=[contact_page.pk],
        )
        response = admin_client.get(url)
        assert response.status_code == 200

    def test_page_title_is_form_data(self, admin_client, contact_page, form_submissions):
        url = reverse(
            "custom_contact_form:list_submissions",
            args=[contact_page.pk],
        )
        response = admin_client.get(url)
        assert response.status_code == 200


@pytest.mark.django_db
class TestCustomFormPagesListView:
    def test_columns_do_not_include_origin(self):
        column_names = [col.name for col in CustomFormPagesListView.columns]
        assert "content_type" not in column_names

    def test_columns_include_title(self):
        column_names = [col.name for col in CustomFormPagesListView.columns]
        assert "title" in column_names

    def test_columns_include_updated(self):
        column_names = [col.name for col in CustomFormPagesListView.columns]
        assert "latest_revision_created_at" in column_names

    def test_filterset_class(self):
        assert CustomFormPagesListView.filterset_class == FormPagesFilterSet

    def test_paginator_class(self):
        assert CustomFormPagesListView.paginator_class == FormPagePaginator

    def test_paginator_verbose_name_plural(self):
        paginator = FormPagePaginator([], per_page=10)
        assert paginator.verbose_name_plural == "Pages"

    def test_paginator_verbose_name(self):
        paginator = FormPagePaginator([], per_page=10)
        assert paginator.verbose_name == "Page"


@pytest.mark.django_db
class TestFormPagesFilterSet:
    def test_filter_label_is_updated(self):
        filterset = FormPagesFilterSet()
        assert filterset.filters["latest_revision_created_at"].label == "Updated"

    def test_only_date_filter_present(self):
        filterset = FormPagesFilterSet()
        assert "latest_revision_created_at" in filterset.filters
        assert "owner" not in filterset.filters
        assert "has_child_pages" not in filterset.filters
        assert "content_type" not in filterset.filters


@pytest.mark.django_db
class TestFormPagesListViewIntegration:
    def test_forms_index_returns_200(self, admin_client, contact_page):
        url = reverse("custom_contact_form:index")
        response = admin_client.get(url)
        assert response.status_code == 200

    def test_forms_index_does_not_show_origin_column(self, admin_client, contact_page):
        url = reverse("custom_contact_form:index")
        response = admin_client.get(url)
        content = response.content.decode("utf-8")
        assert "Origin" not in content
